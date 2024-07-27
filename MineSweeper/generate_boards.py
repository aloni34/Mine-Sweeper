import openpyxl
from openpyxl import load_workbook
import numpy as np
import os


class BoardGenerator(object):

    # region Methods
    @staticmethod
    def generate_boards(amount, h, w, bombs, path, amount_per_excel):

        for i in range(amount // amount_per_excel):  # divides the data into different excels based on amount_per_excel
            for j in range(amount_per_excel):  # amount of columns per excel
                board = BoardGenerator.generate_board(h, w, bombs)  # generate boards
                # generate boards
                BoardGenerator.save_board_to_excel(board, f'{path}/{h}X{w}', f'{h}X{w}-{bombs}-part{i+1}.xlsx')
                print(f"Board {str(j)}")

    @staticmethod
    def generate_board(h, w, bombs):

        # Initialize an empty board
        board = np.zeros((h, w), dtype=int)

        # Flatten the board to a 1D array to make it easier to pick random positions
        flat_board = board.flatten()

        # Randomly select k unique positions to place bombs
        bomb_positions = np.random.choice(h * w, bombs, replace=False)

        # Place bombs (represented by 9) at the selected positions
        flat_board[bomb_positions] = 9

        # Reshape the board back to its original dimensions (in case we want to save in 2D mode)
        # board = flat_board.reshape((h, w))

        return flat_board

    @staticmethod
    def save_board_to_excel(board, directory, filename):

        path = os.path.join(directory, filename)

        try:
            # Load the workbook and select the active worksheet
            workbook = load_workbook(path)
            sheet = workbook.active
        except FileNotFoundError:
            # If the file doesn't exist, create a new workbook and select the active worksheet
            workbook = openpyxl.Workbook()
            sheet = workbook.active

        # Find the first empty column
        max_col = sheet.max_column
        new_col = max_col+1

        # Write data to the new column
        for row_num, value in enumerate(board, start=1):
            sheet.cell(row=row_num, column=new_col, value=value)

        # Save the workbook
        workbook.save(path)
    # endregion


# This might take a while, depends on the amount of boards required and the amount per excel
BoardGenerator.generate_boards(100, 20, 20, 15, 'Data/Boards', 10)
